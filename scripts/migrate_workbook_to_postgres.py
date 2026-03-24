from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from psycopg import sql


ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app_config import AppConfig
from db_config import connect, load_project_env


TABLE_NAME_MAP = {
    AppConfig.SHEET_PARTS: "parts",
    AppConfig.SHEET_MODULES: "modules",
    AppConfig.SHEET_TASKS: "tasks",
    AppConfig.SHEET_COMPONENTS: "components",
    AppConfig.SHEET_DOCUMENTS: "documents",
    AppConfig.SHEET_PRODUCTS: "products",
    AppConfig.SHEET_PRODUCT_MODULES: "product_modules",
    AppConfig.SHEET_PRODUCT_DOCUMENTS: "product_documents",
    AppConfig.SHEET_PROJECTS: "projects",
    AppConfig.SHEET_PROJECT_MODULES: "project_modules",
    AppConfig.SHEET_PROJECT_TASKS: "project_tasks",
    AppConfig.SHEET_PROJECT_DOCUMENTS: "project_documents",
    AppConfig.SHEET_WORKORDERS: "workorders",
    AppConfig.SHEET_COMPLETED_JOBS: "completed_jobs",
    AppConfig.SHEET_COMPLETED_JOB_LINES: "completed_job_lines",
}

COLUMN_NAME_MAP = {
    "PartID": "part_id",
    "PartNumber": "part_number",
    "PartName": "part_name",
    "Description": "description",
    "UnitPrice": "unit_price",
    "LeadTimeDays": "lead_time_days",
    "PreferredSupplier": "preferred_supplier",
    "StockOnHand": "stock_on_hand",
    "Manufacturer": "manufacturer",
    "Category": "category",
    "Notes": "notes",
    "CreatedOn": "created_on",
    "UpdatedOn": "updated_on",
    "ModuleCode": "module_code",
    "QuoteRef": "quote_ref",
    "ModuleName": "module_name",
    "InstructionText": "instruction_text",
    "EstimatedHours": "estimated_hours",
    "Status": "status",
    "TaskID": "task_id",
    "OwnerType": "owner_type",
    "OwnerCode": "owner_code",
    "TaskName": "task_name",
    "Department": "department",
    "ParentTaskID": "parent_task_id",
    "DependencyTaskID": "dependency_task_id",
    "Stage": "stage",
    "ComponentID": "component_id",
    "ComponentName": "component_name",
    "Qty": "qty",
    "SOHQty": "soh_qty",
    "DocID": "doc_id",
    "SectionName": "section_name",
    "DocName": "doc_name",
    "DocType": "doc_type",
    "FilePath": "file_path",
    "AddedOn": "added_on",
    "ProductCode": "product_code",
    "ProductName": "product_name",
    "Revision": "revision",
    "LinkID": "link_id",
    "ModuleOrder": "module_order",
    "ModuleQty": "module_qty",
    "DependencyModuleCode": "dependency_module_code",
    "ProdDocID": "prod_doc_id",
    "ProjectCode": "project_code",
    "ProjectName": "project_name",
    "ClientName": "client_name",
    "Location": "location",
    "LinkedProductCode": "linked_product_code",
    "StartDate": "start_date",
    "DueDate": "due_date",
    "SourceType": "source_type",
    "SourceCode": "source_code",
    "ProjectTaskID": "project_task_id",
    "SourceTaskID": "source_task_id",
    "ParentProjectTaskID": "parent_project_task_id",
    "AssignedTo": "assigned_to",
    "ProjectDocID": "project_doc_id",
    "WorkOrderID": "workorder_id",
    "WorkOrderName": "workorder_name",
    "Owner": "owner",
    "SnapshotID": "snapshot_id",
    "CompletedOn": "completed_on",
    "LabourHours": "labour_hours",
    "PartsTotal": "parts_total",
    "GrandTotal": "grand_total",
    "SnapshotLineID": "snapshot_line_id",
    "LineType": "line_type",
    "Code": "code",
    "Hours": "hours",
    "LineTotal": "line_total",
    "Source": "source",
}

NUMERIC_FIELDS = {
    "UnitPrice",
    "StockOnHand",
    "EstimatedHours",
    "Qty",
    "SOHQty",
    "ModuleQty",
    "LabourHours",
    "PartsTotal",
    "GrandTotal",
    "Hours",
    "LineTotal",
}

INTEGER_FIELDS = {
    "LeadTimeDays",
    "ModuleOrder",
}

UNIT_PRICE_RE = re.compile(r"UnitPrice=([0-9]+(?:\.[0-9]+)?)")


@dataclass
class SheetValidationResult:
    sheet_name: str
    present: bool
    missing_headers: list[str] = field(default_factory=list)
    extra_headers: list[str] = field(default_factory=list)
    reordered_headers: list[str] = field(default_factory=list)
    duplicate_headers: list[str] = field(default_factory=list)

    @property
    def is_blocking(self) -> bool:
        return bool(self.missing_headers or self.duplicate_headers)


@dataclass
class WorkbookValidationReport:
    workbook_path: Path
    missing_sheets: list[str] = field(default_factory=list)
    extra_sheets: list[str] = field(default_factory=list)
    sheet_results: list[SheetValidationResult] = field(default_factory=list)

    @property
    def has_blocking_issues(self) -> bool:
        return any(result.is_blocking for result in self.sheet_results)


def looks_like_number(value: Any) -> bool:
    if value in (None, ""):
        return False
    try:
        float(value)
        return True
    except Exception:
        return False


def looks_like_datetime_text(value: Any) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    if not isinstance(value, str):
        return False
    text = value.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def repair_legacy_row(sheet_name: str, padded_row: list[Any]) -> list[Any]:
    if sheet_name != AppConfig.SHEET_COMPONENTS:
        return padded_row

    repaired = list(padded_row)
    unit_price, part_number, notes, created_on, updated_on = repaired[8:13]

    has_shifted_metadata = (
        not looks_like_number(unit_price)
        and looks_like_datetime_text(part_number)
        and isinstance(notes, str)
    )
    if not has_shifted_metadata:
        return padded_row

    parsed_unit_price = None
    parsed_updated_on = None
    note_text = unit_price if isinstance(unit_price, str) else None

    if isinstance(notes, str):
        parts = [part.strip() for part in notes.split("|")]
        if parts and looks_like_datetime_text(parts[0]):
            parsed_updated_on = parts[0]
        match = UNIT_PRICE_RE.search(notes)
        if match:
            parsed_unit_price = float(match.group(1))

    repaired[8] = parsed_unit_price
    repaired[9] = None
    repaired[10] = note_text

    if created_on in (None, "") and looks_like_datetime_text(part_number):
        repaired[11] = part_number
    else:
        repaired[11] = created_on

    if updated_on in (None, "") and parsed_updated_on:
        repaired[12] = parsed_updated_on
    else:
        repaired[12] = updated_on

    return repaired


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import workbook data into PostgreSQL.")
    parser.add_argument(
        "--workbook",
        default=os.getenv("ERP_WORKBOOK_PATH", str(ROOT_DIR / AppConfig.DEFAULT_WORKBOOK_NAME)),
        help="Path to the workbook to import.",
    )
    parser.add_argument(
        "--schema-file",
        default=str(ROOT_DIR / "sql" / "postgres_schema.sql"),
        help="Path to the SQL schema file.",
    )
    parser.add_argument(
        "--schema-only",
        action="store_true",
        help="Create or update the schema without importing workbook data.",
    )
    parser.add_argument(
        "--skip-schema",
        action="store_true",
        help="Skip applying the schema file before importing.",
    )
    parser.add_argument(
        "--truncate-first",
        action="store_true",
        help="Delete existing rows from the target tables before import.",
    )
    parser.add_argument(
        "--preflight-only",
        action="store_true",
        help="Validate workbook sheets and headers without importing data.",
    )
    return parser.parse_args()


def normalize_cell(value: Any, field_name: str) -> Any:
    if value in ("", None):
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if field_name in INTEGER_FIELDS:
        try:
            return int(float(value))
        except Exception:
            return None
    if field_name in NUMERIC_FIELDS:
        try:
            return float(value)
        except Exception:
            return None
    return str(value)


def non_empty_rows(ws):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(value not in (None, "") for value in row):
            yield list(row)


def apply_schema(conn, schema_path: Path) -> None:
    sql_text = schema_path.read_text(encoding="utf-8")
    with conn.cursor() as cur:
        cur.execute(sql_text)
    conn.commit()


def truncate_tables(conn) -> None:
    with conn.cursor() as cur:
        for table_name in reversed(list(TABLE_NAME_MAP.values())):
            cur.execute(sql.SQL("TRUNCATE TABLE {}").format(sql.Identifier(table_name)))
    conn.commit()


def validate_workbook_schema(workbook_path: Path) -> WorkbookValidationReport:
    wb = load_workbook(workbook_path, data_only=True)
    report = WorkbookValidationReport(workbook_path=workbook_path)
    report.missing_sheets = [sheet for sheet in AppConfig.ALL_SHEETS if sheet not in wb.sheetnames]
    report.extra_sheets = [sheet for sheet in wb.sheetnames if sheet not in AppConfig.ALL_SHEETS]

    for sheet_name in AppConfig.ALL_SHEETS:
        if sheet_name not in wb.sheetnames:
            report.sheet_results.append(SheetValidationResult(sheet_name=sheet_name, present=False))
            continue

        expected_headers = AppConfig.SHEET_HEADERS[sheet_name]
        ws = wb[sheet_name]
        actual_headers = [header for header in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        actual_positions: dict[str, int] = {}
        duplicates: list[str] = []
        for idx, header in enumerate(actual_headers):
            if header in (None, ""):
                continue
            header_text = str(header)
            if header_text in actual_positions and header_text not in duplicates:
                duplicates.append(header_text)
            actual_positions.setdefault(header_text, idx)

        missing_headers = [header for header in expected_headers if header not in actual_positions]
        extra_headers = [str(header) for header in actual_headers if header not in (None, "") and str(header) not in expected_headers]
        reordered_headers = [
            header
            for expected_idx, header in enumerate(expected_headers)
            if header in actual_positions and actual_positions[header] != expected_idx
        ]

        report.sheet_results.append(
            SheetValidationResult(
                sheet_name=sheet_name,
                present=True,
                missing_headers=missing_headers,
                extra_headers=extra_headers,
                reordered_headers=reordered_headers,
                duplicate_headers=duplicates,
            )
        )

    return report


def print_validation_report(report: WorkbookValidationReport) -> None:
    print(f"Workbook schema preflight: {report.workbook_path}")
    if report.extra_sheets:
        print("Extra workbook sheets:")
        for sheet_name in report.extra_sheets:
            print(f"- {sheet_name}")
    if report.missing_sheets:
        print("Missing expected sheets:")
        for sheet_name in report.missing_sheets:
            print(f"- {sheet_name}")

    for result in report.sheet_results:
        if not result.present:
            continue
        if not (result.missing_headers or result.extra_headers or result.reordered_headers or result.duplicate_headers):
            continue
        print(f"Sheet: {result.sheet_name}")
        if result.missing_headers:
            print(f"- Missing headers: {', '.join(result.missing_headers)}")
        if result.extra_headers:
            print(f"- Extra headers: {', '.join(result.extra_headers)}")
        if result.reordered_headers:
            print(f"- Reordered headers: {', '.join(result.reordered_headers)}")
        if result.duplicate_headers:
            print(f"- Duplicate headers: {', '.join(result.duplicate_headers)}")

    if report.has_blocking_issues:
        print("Preflight result: blocking schema issues found.")
    else:
        print("Preflight result: workbook is compatible for import.")


def import_sheet(conn, ws, sheet_name: str) -> int:
    headers = AppConfig.SHEET_HEADERS[sheet_name]
    actual_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    header_positions: dict[str, int] = {}
    duplicates: list[str] = []
    for idx, header in enumerate(actual_headers):
        if header in (None, ""):
            continue
        header_text = str(header)
        if header_text in header_positions and header_text not in duplicates:
            duplicates.append(header_text)
        header_positions.setdefault(header_text, idx)

    missing_headers = [header for header in headers if header not in header_positions]
    if missing_headers or duplicates:
        problems: list[str] = []
        if missing_headers:
            problems.append(f"missing headers: {', '.join(missing_headers)}")
        if duplicates:
            problems.append(f"duplicate headers: {', '.join(duplicates)}")
        raise ValueError(f"Header mismatch in sheet '{sheet_name}': {'; '.join(problems)}")

    table_name = TABLE_NAME_MAP[sheet_name]
    column_names = [COLUMN_NAME_MAP[header] for header in headers]
    primary_key_column = column_names[0]
    update_columns = column_names[1:]

    insert_stmt = sql.SQL(
        """
        INSERT INTO {table_name} ({columns})
        VALUES ({values})
        ON CONFLICT ({primary_key})
        DO UPDATE SET {updates}
        """
    ).format(
        table_name=sql.Identifier(table_name),
        columns=sql.SQL(", ").join(sql.Identifier(col) for col in column_names),
        values=sql.SQL(", ").join(sql.Placeholder() for _ in column_names),
        primary_key=sql.Identifier(primary_key_column),
        updates=sql.SQL(", ").join(
            sql.SQL("{col} = EXCLUDED.{col}").format(col=sql.Identifier(col))
            for col in update_columns
        ),
    )

    imported = 0
    with conn.cursor() as cur:
        for raw_row in non_empty_rows(ws):
            actual_row = list(raw_row)
            padded_row = [
                actual_row[header_positions[header]] if header_positions[header] < len(actual_row) else None
                for header in headers
            ]
            padded_row = repair_legacy_row(sheet_name, padded_row)
            values = [normalize_cell(value, header) for header, value in zip(headers, padded_row)]
            if values[0] in (None, ""):
                continue
            cur.execute(insert_stmt, values)
            imported += 1
    return imported


def main() -> int:
    args = parse_args()
    env_path = load_project_env()
    workbook_path = Path(args.workbook).expanduser().resolve()
    schema_path = Path(args.schema_file).expanduser().resolve()

    if not args.skip_schema and not schema_path.exists():
        print(f"Schema file not found: {schema_path}")
        return 1

    if not args.schema_only and not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        return 1

    print(f"Loaded environment from: {env_path}")
    if not args.schema_only or args.preflight_only:
        print(f"Workbook: {workbook_path}")

    if not args.schema_only or args.preflight_only:
        report = validate_workbook_schema(workbook_path)
        print_validation_report(report)
        if report.has_blocking_issues:
            return 1
        if args.preflight_only:
            return 0

    with connect() as conn:
        if not args.skip_schema:
            print(f"Applying schema from: {schema_path}")
            apply_schema(conn, schema_path)

        if args.schema_only:
            print("Schema applied successfully. Data import skipped.")
            return 0

        if args.truncate_first:
            print("Truncating existing table data before import...")
            truncate_tables(conn)

        wb = load_workbook(workbook_path, data_only=True)
        counts: list[tuple[str, int]] = []
        missing_sheets: list[str] = []
        for sheet_name in AppConfig.ALL_SHEETS:
            if sheet_name not in wb.sheetnames:
                missing_sheets.append(sheet_name)
                print(f"Skipping missing sheet: {sheet_name}")
                continue
            ws = wb[sheet_name]
            count = import_sheet(conn, ws, sheet_name)
            counts.append((sheet_name, count))
            print(f"Imported {count} row(s) from {sheet_name}.")

        conn.commit()

    print("\nImport complete.")
    for sheet_name, count in counts:
        print(f"- {sheet_name}: {count}")
    if missing_sheets:
        print("Skipped missing sheets:")
        for sheet_name in missing_sheets:
            print(f"- {sheet_name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
