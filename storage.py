# # ============================================================
# # storage.py
# # Excel storage and persistence layer for Liquimech ERP Desktop
# # ============================================================

# from __future__ import annotations

# from pathlib import Path
# from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment

# from app_config import AppConfig


# # ============================================================
# # Basic helpers
# # ============================================================

# def now_str() -> str:
#     from datetime import datetime
#     return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# def norm_text(value: Any) -> str:
#     return str(value or "").strip()


# def to_float(value: Any, default: float = 0.0) -> float:
#     try:
#         if value in (None, ""):
#             return default
#         return float(value)
#     except Exception:
#         return default


# def to_int(value: Any, default: int = 0) -> int:
#     try:
#         if value in (None, ""):
#             return default
#         return int(float(value))
#     except Exception:
#         return default


# def safe_name(text: str) -> str:
#     return str(text or "").strip().replace("/", "_").replace("\\", "_")


# # ============================================================
# # Workbook schema / styling
# # ============================================================

# class WorkbookSchema:
#     @staticmethod
#     def style_sheet(ws) -> None:
#         header_fill = PatternFill("solid", fgColor="1F4E78")
#         header_font = Font(color="FFFFFF", bold=True)

#         for cell in ws[1]:
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#         ws.freeze_panes = "A2"

#         widths = {
#             "A": 24, "B": 22, "C": 28, "D": 32, "E": 18,
#             "F": 18, "G": 20, "H": 24, "I": 18, "J": 18,
#             "K": 18, "L": 18, "M": 18, "N": 18, "O": 18
#         }
#         for col, width in widths.items():
#             ws.column_dimensions[col].width = width

#     @classmethod
#     def create_new_workbook(cls, path: str | Path) -> None:
#         path = str(path)
#         wb = Workbook()

#         first_sheet_name = AppConfig.ALL_SHEETS[0]
#         ws = wb.active
#         ws.title = first_sheet_name
#         ws.append(AppConfig.SHEET_HEADERS[first_sheet_name])
#         cls.style_sheet(ws)

#         for sheet_name in AppConfig.ALL_SHEETS[1:]:
#             new_ws = wb.create_sheet(sheet_name)
#             new_ws.append(AppConfig.SHEET_HEADERS[sheet_name])
#             cls.style_sheet(new_ws)

#         wb.save(path)

#     @classmethod
#     def ensure_workbook_structure(cls, path: str | Path) -> None:
#         path = str(path)
#         p = Path(path)

#         if not p.exists():
#             cls.create_new_workbook(path)
#             return

#         wb = load_workbook(path)
#         changed = False

#         for sheet_name in AppConfig.ALL_SHEETS:
#             headers = AppConfig.SHEET_HEADERS[sheet_name]

#             if sheet_name not in wb.sheetnames:
#                 ws = wb.create_sheet(sheet_name)
#                 ws.append(headers)
#                 cls.style_sheet(ws)
#                 changed = True
#                 continue

#             ws = wb[sheet_name]
#             existing_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(headers))]
#             if existing_headers != headers:
#                 for i, h in enumerate(headers, start=1):
#                     ws.cell(row=1, column=i, value=h)
#                 cls.style_sheet(ws)
#                 changed = True

#         if changed:
#             wb.save(path)


# # ============================================================
# # Repository
# # ============================================================

# class ExcelRepository:
#     """
#     Thin repository over an Excel workbook.

#     Rules:
#     - UI should not directly mutate workbook rows everywhere.
#     - Services should go through this class.
#     - Saves happen immediately after change operations.
#     """

#     def __init__(self, workbook_path: Optional[str] = None):
#         self._sheet_cache = {}
#         self.workbook_path: Optional[str] = workbook_path
#         self._wb_cache = None
#         self._dirty = False

#     # --------------------------------------------------------
#     # Workbook file handling
#     # --------------------------------------------------------

#     def invalidate_sheet_cache(self, sheet_name: Optional[str] = None):
#         if sheet_name is None:
#             self._sheet_cache = {}
#         else:
#             self._sheet_cache.pop(sheet_name, None)


#     def get_cached_rows(self, sheet_name: str):
#         if sheet_name in self._sheet_cache:
#             return self._sheet_cache[sheet_name]

#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)
#         rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
#         self._sheet_cache[sheet_name] = rows
#         return rows
    
#     def set_workbook_path(self, workbook_path: str) -> None:
#         self.workbook_path = workbook_path
#         self._wb_cache = None
#         self._dirty = False

#     def has_workbook(self) -> bool:
#         return bool(self.workbook_path)

#     def require_workbook_path(self) -> str:
#         if not self.workbook_path:
#             raise ValueError("No workbook selected.")
#         return self.workbook_path

#     def ensure_ready(self) -> None:
#         workbook_path = self.require_workbook_path()
#         AppConfig.ensure_directories()
#         WorkbookSchema.ensure_workbook_structure(workbook_path)

#     # def load_workbook_safe(self):
#     #     workbook_path = self.require_workbook_path()
#     #     WorkbookSchema.ensure_workbook_structure(workbook_path)
#     #     return load_workbook(workbook_path)

#     def load_workbook_safe(self):
#         workbook_path = self.require_workbook_path()
#         WorkbookSchema.ensure_workbook_structure(workbook_path)

#         if self._wb_cache is None:
#             self._wb_cache = load_workbook(workbook_path)

#         return self._wb_cache

#     # def save_workbook(self, wb) -> None:
#     #     workbook_path = self.require_workbook_path()
#     #     wb.save(workbook_path)
#     def save_workbook(self, wb=None) -> None:
#         workbook_path = self.require_workbook_path()

#         if wb is None:
#             wb = self._wb_cache

#         if wb is None:
#             return

#         wb.save(workbook_path)
#         self._dirty = False

#     def mark_dirty(self) -> None:
#         self._dirty = True

#     def mark_dirty(self) -> None:
#         self._dirty = True
#     # --------------------------------------------------------
#     # Sheet helpers
#     # --------------------------------------------------------
#     def get_sheet_headers(self, sheet_name: str) -> List[str]:
#         headers = AppConfig.SHEET_HEADERS.get(sheet_name)
#         if not headers:
#             raise ValueError(f"Unknown sheet name: {sheet_name}")
#         return headers

#     def get_sheet(self, wb, sheet_name: str):
#         if sheet_name not in wb.sheetnames:
#             raise ValueError(f"Sheet not found: {sheet_name}")
#         return wb[sheet_name]

#     def get_header_index_map(self, sheet_name: str) -> Dict[str, int]:
#         headers = self.get_sheet_headers(sheet_name)
#         return {h: i for i, h in enumerate(headers)}

#     # --------------------------------------------------------
#     # Generic row reading
#     # --------------------------------------------------------
#     def list_rows(self, sheet_name: str) -> List[List[Any]]:
#         self.ensure_ready()
#         return self.get_cached_rows(sheet_name)
#         # wb = self.load_workbook_safe()
#         # ws = self.get_sheet(wb, sheet_name)
#         # rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
#         # return rows

#     def list_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
#         headers = self.get_sheet_headers(sheet_name)
#         output = []
#         for row in self.list_rows(sheet_name):
#             padded = list(row) + [None] * (len(headers) - len(row))
#             output.append(dict(zip(headers, padded)))
#         return output

#     def filter_rows(self, sheet_name: str, predicate: Callable[[List[Any]], bool]) -> List[List[Any]]:
#         return [row for row in self.list_rows(sheet_name) if predicate(row)]

#     def filter_dicts(self, sheet_name: str, predicate: Callable[[Dict[str, Any]], bool]) -> List[Dict[str, Any]]:
#         return [row for row in self.list_dicts(sheet_name) if predicate(row)]

#     def find_row(self, sheet_name: str, key_col_idx: int, key_value: Any) -> Optional[List[Any]]:
#         key_value = norm_text(key_value)
#         for row in self.list_rows(sheet_name):
#             if key_col_idx < len(row) and norm_text(row[key_col_idx]) == key_value:
#                 return row
#         return None

#     def find_row_index(self, sheet_name: str, key_col_idx: int, key_value: Any) -> Optional[int]:
#         """
#         Returns actual Excel row number (starting at 2 for first data row)
#         """
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)

#         key_value = norm_text(key_value)
#         for excel_row in range(2, ws.max_row + 1):
#             if norm_text(ws.cell(excel_row, key_col_idx + 1).value) == key_value:
#                 return excel_row
#         return None

#     def exists(self, sheet_name: str, key_col_idx: int, key_value: Any) -> bool:
#         return self.find_row_index(sheet_name, key_col_idx, key_value) is not None

#     # --------------------------------------------------------
#     # Append / insert
#     # --------------------------------------------------------
#     def append_row(self, sheet_name: str, values: Sequence[Any]) -> None:
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)
#         ws.append(list(values))
#         self.mark_dirty()
#         self.save_workbook(wb)

#     def append_dict(self, sheet_name: str, row_dict: Dict[str, Any]) -> None:
#         headers = self.get_sheet_headers(sheet_name)
#         row = [row_dict.get(h, "") for h in headers]
#         self.append_row(sheet_name, row)

#     # --------------------------------------------------------
#     # Update operations
#     # --------------------------------------------------------
#     def update_row_by_key(
#         self,
#         sheet_name: str,
#         key_col_idx: int,
#         key_value: Any,
#         updates: Dict[int, Any]
#     ) -> bool:
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)

#         key_value = norm_text(key_value)
#         for excel_row in range(2, ws.max_row + 1):
#             if norm_text(ws.cell(excel_row, key_col_idx + 1).value) == key_value:
#                 for col_idx, value in updates.items():
#                     ws.cell(excel_row, col_idx + 1, value)
#                 self.save_workbook(wb)
#                 return True
#         return False

#     def update_row_by_key_name(
#         self,
#         sheet_name: str,
#         key_name: str,
#         key_value: Any,
#         updates: Dict[str, Any]
#     ) -> bool:
#         header_map = self.get_header_index_map(sheet_name)
#         if key_name not in header_map:
#             raise ValueError(f"Unknown key column '{key_name}' for sheet '{sheet_name}'")

#         converted_updates = {}
#         for field_name, value in updates.items():
#             if field_name not in header_map:
#                 continue
#             converted_updates[header_map[field_name]] = value

#         return self.update_row_by_key(
#             sheet_name=sheet_name,
#             key_col_idx=header_map[key_name],
#             key_value=key_value,
#             updates=converted_updates
#         )

#     def replace_full_row_by_key(
#         self,
#         sheet_name: str,
#         key_col_idx: int,
#         key_value: Any,
#         new_row: Sequence[Any]
#     ) -> bool:
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)

#         headers = self.get_sheet_headers(sheet_name)
#         row_values = list(new_row)[:len(headers)]
#         row_values += [""] * (len(headers) - len(row_values))

#         key_value = norm_text(key_value)
#         for excel_row in range(2, ws.max_row + 1):
#             if norm_text(ws.cell(excel_row, key_col_idx + 1).value) == key_value:
#                 for i, value in enumerate(row_values, start=1):
#                     ws.cell(excel_row, i, value)
#                 self.save_workbook(wb)
#                 return True
#         return False

#     def upsert_row(
#         self,
#         sheet_name: str,
#         key_col_idx: int,
#         key_value: Any,
#         full_row: Sequence[Any]
#     ) -> str:
#         """
#         Returns 'updated' or 'inserted'
#         """
#         if self.exists(sheet_name, key_col_idx, key_value):
#             self.replace_full_row_by_key(sheet_name, key_col_idx, key_value, full_row)
#             return "updated"

#         self.append_row(sheet_name, full_row)
#         return "inserted"

#     def upsert_dict(
#         self,
#         sheet_name: str,
#         key_name: str,
#         row_dict: Dict[str, Any]
#     ) -> str:
#         headers = self.get_sheet_headers(sheet_name)
#         header_map = self.get_header_index_map(sheet_name)
#         if key_name not in header_map:
#             raise ValueError(f"Unknown key column '{key_name}' for sheet '{sheet_name}'")

#         key_value = row_dict.get(key_name, "")
#         row = [row_dict.get(h, "") for h in headers]
#         return self.upsert_row(sheet_name, header_map[key_name], key_value, row)

#     # --------------------------------------------------------
#     # Delete operations
#     # --------------------------------------------------------
#     def delete_row_by_key(self, sheet_name: str, key_col_idx: int, key_value: Any) -> bool:
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)

#         key_value = norm_text(key_value)
#         for excel_row in range(2, ws.max_row + 1):
#             if norm_text(ws.cell(excel_row, key_col_idx + 1).value) == key_value:
#                 ws.delete_rows(excel_row, 1)
#                 self.save_workbook(wb)
#                 return True
#         return False

#     def delete_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any) -> bool:
#         header_map = self.get_header_index_map(sheet_name)
#         if key_name not in header_map:
#             raise ValueError(f"Unknown key column '{key_name}' for sheet '{sheet_name}'")
#         return self.delete_row_by_key(sheet_name, header_map[key_name], key_value)

#     def delete_rows_where(self, sheet_name: str, predicate: Callable[[Dict[str, Any]], bool]) -> int:
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)
#         headers = self.get_sheet_headers(sheet_name)

#         rows_to_delete = []
#         for excel_row in range(2, ws.max_row + 1):
#             row_vals = [ws.cell(excel_row, i + 1).value for i in range(len(headers))]
#             row_dict = dict(zip(headers, row_vals))
#             if predicate(row_dict):
#                 rows_to_delete.append(excel_row)

#         for excel_row in reversed(rows_to_delete):
#             ws.delete_rows(excel_row, 1)

#         if rows_to_delete:
#             self.save_workbook(wb)

#         return len(rows_to_delete)

#     # --------------------------------------------------------
#     # Bulk / utility operations
#     # --------------------------------------------------------
#     def clear_sheet_data(self, sheet_name: str) -> None:
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)

#         if ws.max_row >= 2:
#             ws.delete_rows(2, ws.max_row - 1)

#         self.save_workbook(wb)

#     def rewrite_sheet_data(self, sheet_name: str, rows: List[Sequence[Any]]) -> None:
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)
#         headers = self.get_sheet_headers(sheet_name)

#         if ws.max_row >= 2:
#             ws.delete_rows(2, ws.max_row - 1)

#         for row in rows:
#             row_vals = list(row)[:len(headers)]
#             row_vals += [""] * (len(headers) - len(row_vals))
#             ws.append(row_vals)

#         self.save_workbook(wb)

#     def reorder_rows_by_field(
#         self,
#         sheet_name: str,
#         filter_key: str,
#         filter_value: Any,
#         order_field: str,
#         ordered_key_field: str,
#         ordered_key_values: List[Any]
#     ) -> None:
#         """
#         Reorders a filtered group of records by assigning order_field = 1..N
#         based on ordered_key_values.

#         Example:
#         reorder_rows_by_field(
#             sheet_name="ProductModules",
#             filter_key="ProductCode",
#             filter_value="P_ABC",
#             order_field="ModuleOrder",
#             ordered_key_field="ModuleCode",
#             ordered_key_values=["M1", "M2", "M3"]
#         )
#         """
#         self.ensure_ready()
#         wb = self.load_workbook_safe()
#         ws = self.get_sheet(wb, sheet_name)
#         header_map = self.get_header_index_map(sheet_name)

#         if filter_key not in header_map or order_field not in header_map or ordered_key_field not in header_map:
#             raise ValueError("Invalid field name in reorder_rows_by_field")

#         filter_idx = header_map[filter_key]
#         order_idx = header_map[order_field]
#         ordered_key_idx = header_map[ordered_key_field]

#         order_lookup = {norm_text(v): i + 1 for i, v in enumerate(ordered_key_values)}

#         changed = False
#         for excel_row in range(2, ws.max_row + 1):
#             if norm_text(ws.cell(excel_row, filter_idx + 1).value) != norm_text(filter_value):
#                 continue

#             key_val = norm_text(ws.cell(excel_row, ordered_key_idx + 1).value)
#             if key_val in order_lookup:
#                 ws.cell(excel_row, order_idx + 1, order_lookup[key_val])
#                 changed = True

#         if changed:
#             self.save_workbook(wb)

#     # --------------------------------------------------------
#     # Relationship helpers
#     # --------------------------------------------------------
#     def get_rows_by_owner(self, sheet_name: str, owner_type: str, owner_code: str) -> List[Dict[str, Any]]:
#         return self.filter_dicts(
#             sheet_name,
#             lambda r: norm_text(r.get("OwnerType")) == norm_text(owner_type)
#             and norm_text(r.get("OwnerCode")) == norm_text(owner_code)
#         )

#     def delete_rows_by_owner(self, sheet_name: str, owner_type: str, owner_code: str) -> int:
#         return self.delete_rows_where(
#             sheet_name,
#             lambda r: norm_text(r.get("OwnerType")) == norm_text(owner_type)
#             and norm_text(r.get("OwnerCode")) == norm_text(owner_code)
#         )

#     # --------------------------------------------------------
#     # Docs folder helpers
#     # --------------------------------------------------------
#     def get_module_docs_folder(self, module_code: str) -> Path:
#         AppConfig.ensure_directories()
#         folder = AppConfig.MODULE_DOCS_DIR / safe_name(module_code)
#         folder.mkdir(parents=True, exist_ok=True)
#         return folder

#     def get_product_docs_folder(self, product_code: str) -> Path:
#         AppConfig.ensure_directories()
#         folder = AppConfig.PRODUCT_DOCS_DIR / safe_name(product_code)
#         folder.mkdir(parents=True, exist_ok=True)
#         return folder

#     def get_project_docs_folder(self, project_code: str) -> Path:
#         AppConfig.ensure_directories()
#         folder = AppConfig.PROJECT_DOCS_DIR / safe_name(project_code)
#         folder.mkdir(parents=True, exist_ok=True)
#         return folder


# # ============================================================
# # Convenience facade
# # ============================================================

# class WorkbookManager:
#     """
#     Small helper wrapper around ExcelRepository for UI code.

#     Lets pages do simple checks like:
#     - has workbook?
#     - create workbook
#     - open workbook
#     """

#     def __init__(self):
#         self.repo = ExcelRepository()

#     @property
#     def workbook_path(self) -> Optional[str]:
#         return self.repo.workbook_path

#     def set_workbook_path(self, path: str) -> None:
#         self.repo.set_workbook_path(path)

#     def has_workbook(self) -> bool:
#         return self.repo.has_workbook()

#     def create_workbook(self, path: str) -> None:
#         AppConfig.ensure_directories()
#         WorkbookSchema.create_new_workbook(path)
#         self.repo.set_workbook_path(path)

#     def open_workbook(self, path: str) -> None:
#         AppConfig.ensure_directories()
#         WorkbookSchema.ensure_workbook_structure(path)
#         self.repo.set_workbook_path(path)

# ============================================================
# storage.py
# Excel storage and persistence layer for Liquimech ERP Desktop
# ============================================================

from __future__ import annotations

import base64
import json
import mimetypes
import os
from contextlib import contextmanager
from importlib import import_module
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence
from urllib.parse import urlencode
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app_config import AppConfig


# ============================================================
# Basic helpers
# ============================================================

def load_project_env(env_file: Path | None = None) -> Path:
    db_config = import_module("db_config")
    return db_config.load_project_env(env_file)


def get_database_settings() -> dict[str, str]:
    db_config = import_module("db_config")
    return db_config.get_database_settings()


def connect_postgres(**overrides):
    db_config = import_module("db_config")
    return db_config.connect(**overrides)


def get_psycopg_sql():
    return import_module("psycopg.sql")

def now_str() -> str:
    from datetime import datetime
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm_text(value: Any) -> str:
    return str(value or "").strip()


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, ""):
            return default
        return float(value)
    except Exception:
        return default


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(float(value))
    except Exception:
        return default


def safe_name(text: str) -> str:
    return str(text or "").strip().replace("/", "_").replace("\\", "_")


SHEET_TABLE_NAMES = {
    AppConfig.SHEET_PARTS: "parts",
    AppConfig.SHEET_MODULES: "modules",
    AppConfig.SHEET_TASKS: "tasks",
    AppConfig.SHEET_COMPONENTS: "components",
    AppConfig.SHEET_DOCUMENTS: "documents",
    AppConfig.SHEET_PRODUCTS: "products",
    AppConfig.SHEET_PRODUCT_MODULES: "product_modules",
    AppConfig.SHEET_PRODUCT_DOCUMENTS: "product_documents",
    AppConfig.SHEET_PROJECTS: "projects",
    AppConfig.SHEET_PROJECT_MODULES: "project_modules",
    AppConfig.SHEET_PROJECT_TASKS: "project_tasks",
    AppConfig.SHEET_PROJECT_DOCUMENTS: "project_documents",
    AppConfig.SHEET_WORKORDERS: "workorders",
    AppConfig.SHEET_COMPLETED_JOBS: "completed_jobs",
    AppConfig.SHEET_COMPLETED_JOB_LINES: "completed_job_lines",
}

HEADER_COLUMN_OVERRIDES = {
    "WorkOrderID": "workorder_id",
    "WorkOrderName": "workorder_name",
}


def header_to_column_name(header: str) -> str:
    if header in HEADER_COLUMN_OVERRIDES:
        return HEADER_COLUMN_OVERRIDES[header]
    chars: list[str] = []
    for idx, ch in enumerate(header):
        if ch.isupper() and idx > 0 and (header[idx - 1].islower() or (idx + 1 < len(header) and header[idx + 1].islower())):
            chars.append("_")
        chars.append(ch.lower())
    return "".join(chars)


# ============================================================
# Workbook schema / styling
# ============================================================

class WorkbookSchema:
    @staticmethod
    def style_sheet(ws) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        widths = {
            "A": 24, "B": 22, "C": 28, "D": 32, "E": 18,
            "F": 18, "G": 20, "H": 24, "I": 18, "J": 18,
            "K": 18, "L": 18, "M": 18, "N": 18, "O": 18
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    @classmethod
    def create_new_workbook(cls, path: str | Path) -> None:
        path = str(path)
        wb = Workbook()

        first_sheet_name = AppConfig.ALL_SHEETS[0]
        ws = wb.active
        ws.title = first_sheet_name
        ws.append(AppConfig.SHEET_HEADERS[first_sheet_name])
        cls.style_sheet(ws)

        for sheet_name in AppConfig.ALL_SHEETS[1:]:
            new_ws = wb.create_sheet(sheet_name)
            new_ws.append(AppConfig.SHEET_HEADERS[sheet_name])
            cls.style_sheet(new_ws)

        wb.save(path)

    @classmethod
    def ensure_workbook_structure(cls, path: str | Path) -> None:
        path = str(path)
        p = Path(path)

        if not p.exists():
            cls.create_new_workbook(path)
            return

        wb = load_workbook(path)
        changed = False

        for sheet_name in AppConfig.ALL_SHEETS:
            headers = AppConfig.SHEET_HEADERS[sheet_name]

            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)
                cls.style_sheet(ws)
                changed = True
                continue

            ws = wb[sheet_name]
            existing_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(headers))]
            if existing_headers != headers:
                for i, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=i, value=h)
                cls.style_sheet(ws)
                changed = True

        if changed:
            wb.save(path)


# ============================================================
# Repository
# ============================================================

class ExcelRepository:
    """
    Thin repository over an Excel workbook.

    Rules:
    - UI should not directly mutate workbook rows everywhere.
    - Services should go through this class.
    - Workbook is cached in memory for speed.
    - Sheet row data is cached and invalidated on writes.
    """

    def __init__(self, workbook_path: Optional[str] = None):
        self.workbook_path: Optional[str] = workbook_path
        self._wb_cache = None
        self._sheet_cache: Dict[str, List[List[Any]]] = {}
        self._dirty = False
        self._batch_depth = 0
        self._pending_save = False
        self._validated_workbook_path: Optional[str] = None

    # --------------------------------------------------------
    # Cache helpers
    # --------------------------------------------------------
    def invalidate_sheet_cache(self, sheet_name: Optional[str] = None) -> None:
        if sheet_name is None:
            self._sheet_cache = {}
        else:
            self._sheet_cache.pop(sheet_name, None)

    def get_cached_rows(self, sheet_name: str) -> List[List[Any]]:
        if sheet_name in self._sheet_cache:
            return self._sheet_cache[sheet_name]

        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)
        rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        self._sheet_cache[sheet_name] = rows
        return rows

    def reload_cache(self) -> None:
        self._wb_cache = None
        self.invalidate_sheet_cache()
        self._validated_workbook_path = None

    def mark_dirty(self) -> None:
        self._dirty = True

    @contextmanager
    def batch_update(self):
        self.ensure_ready()
        self._batch_depth += 1
        try:
            yield
        finally:
            self._batch_depth -= 1
            if self._batch_depth == 0 and self._pending_save:
                self.save_workbook()
                self._pending_save = False

    # --------------------------------------------------------
    # Workbook file handling
    # --------------------------------------------------------
    def set_workbook_path(self, workbook_path: str) -> None:
        self.workbook_path = workbook_path
        self._wb_cache = None
        self._sheet_cache = {}
        self._dirty = False
        self._pending_save = False
        self._batch_depth = 0
        self._validated_workbook_path = None

    def has_workbook(self) -> bool:
        return bool(self.workbook_path)

    def require_workbook_path(self) -> str:
        if not self.workbook_path:
            raise ValueError("No workbook selected.")
        return self.workbook_path

    def ensure_ready(self) -> None:
        workbook_path = self.require_workbook_path()
        AppConfig.ensure_directories()
        if self._validated_workbook_path != workbook_path:
            WorkbookSchema.ensure_workbook_structure(workbook_path)
            self._validated_workbook_path = workbook_path

    def load_workbook_safe(self):
        workbook_path = self.require_workbook_path()
        self.ensure_ready()

        if self._wb_cache is None:
            self._wb_cache = load_workbook(workbook_path)

        return self._wb_cache

    def save_workbook(self, wb=None) -> None:
        workbook_path = self.require_workbook_path()

        if wb is None:
            wb = self._wb_cache

        if wb is None:
            return

        wb.save(workbook_path)
        self._dirty = False
        self._pending_save = False

    def _save_or_defer(self) -> None:
        if self._batch_depth > 0:
            self._pending_save = True
            return
        self.save_workbook()

    # --------------------------------------------------------
    # Sheet helpers
    # --------------------------------------------------------
    def get_sheet_headers(self, sheet_name: str) -> List[str]:
        headers = AppConfig.SHEET_HEADERS.get(sheet_name)
        if not headers:
            raise ValueError(f"Unknown sheet name: {sheet_name}")
        return headers

    def get_sheet(self, wb, sheet_name: str):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return wb[sheet_name]

    def get_header_index_map(self, sheet_name: str) -> Dict[str, int]:
        headers = self.get_sheet_headers(sheet_name)
        return {h: i for i, h in enumerate(headers)}

    # --------------------------------------------------------
    # Generic row reading
    # --------------------------------------------------------
    def list_rows(self, sheet_name: str) -> List[List[Any]]:
        self.ensure_ready()
        return self.get_cached_rows(sheet_name)

    def list_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
        headers = self.get_sheet_headers(sheet_name)
        output = []
        for row in self.list_rows(sheet_name):
            padded = list(row) + [None] * (len(headers) - len(row))
            output.append(dict(zip(headers, padded)))
        return output

    def filter_rows(self, sheet_name: str, predicate: Callable[[List[Any]], bool]) -> List[List[Any]]:
        return [row for row in self.list_rows(sheet_name) if predicate(row)]

    def filter_dicts(self, sheet_name: str, predicate: Callable[[Dict[str, Any]], bool]) -> List[Dict[str, Any]]:
        return [row for row in self.list_dicts(sheet_name) if predicate(row)]

    def find_row(self, sheet_name: str, key_col_idx: int, key_value: Any) -> Optional[List[Any]]:
        key_value = norm_text(key_value)
        for row in self.list_rows(sheet_name):
            if key_col_idx < len(row) and norm_text(row[key_col_idx]) == key_value:
                return row
        return None

    # --------------------------------------------------------
    # Compatibility aliases for patched service/UI layers
    # --------------------------------------------------------
    def read_sheet_as_rows(self, sheet_name: str) -> List[List[Any]]:
        return self.list_rows(sheet_name)

    def read_sheet_as_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
        return self.list_dicts(sheet_name)

    def find_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any) -> Optional[Dict[str, Any]]:
        header_map = self.get_header_index_map(sheet_name)
        if key_name not in header_map:
            raise ValueError(f"Unknown key column '{key_name}' for sheet '{sheet_name}'")
        row = self.find_row(sheet_name, header_map[key_name], key_value)
        if row is None:
            return None
        headers = self.get_sheet_headers(sheet_name)
        padded = list(row) + [None] * (len(headers) - len(row))
        return dict(zip(headers, padded))

    def find_row_index(self, sheet_name: str, key_col_idx: int, key_value: Any) -> Optional[int]:
        """
        Returns actual Excel row number (starting at 2 for first data row)
        """
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)

        key_value = norm_text(key_value)
        for excel_row in range(2, ws.max_row + 1):
            if norm_text(ws.cell(excel_row, key_col_idx + 1).value) == key_value:
                return excel_row
        return None

    def exists(self, sheet_name: str, key_col_idx: int, key_value: Any) -> bool:
        return self.find_row_index(sheet_name, key_col_idx, key_value) is not None

    # --------------------------------------------------------
    # Append / insert
    # --------------------------------------------------------
    def append_row(self, sheet_name: str, values: Sequence[Any]) -> None:
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)
        ws.append(list(values))
        self.mark_dirty()
        self.invalidate_sheet_cache(sheet_name)
        self._save_or_defer()

    def append_dict(self, sheet_name: str, row_dict: Dict[str, Any]) -> None:
        headers = self.get_sheet_headers(sheet_name)
        row = [row_dict.get(h, "") for h in headers]
        self.append_row(sheet_name, row)

    # --------------------------------------------------------
    # Update operations
    # --------------------------------------------------------
    def update_row_by_key(
        self,
        sheet_name: str,
        key_col_idx: int,
        key_value: Any,
        updates: Dict[int, Any]
    ) -> bool:
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)

        key_value = norm_text(key_value)
        for excel_row in range(2, ws.max_row + 1):
            if norm_text(ws.cell(excel_row, key_col_idx + 1).value) == key_value:
                for col_idx, value in updates.items():
                    ws.cell(excel_row, col_idx + 1, value)
                self.mark_dirty()
                self.invalidate_sheet_cache(sheet_name)
                self._save_or_defer()
                return True
        return False

    def update_row_by_key_name(
        self,
        sheet_name: str,
        key_name: str,
        key_value: Any,
        updates: Dict[str, Any]
    ) -> bool:
        header_map = self.get_header_index_map(sheet_name)
        if key_name not in header_map:
            raise ValueError(f"Unknown key column '{key_name}' for sheet '{sheet_name}'")

        converted_updates = {}
        for field_name, value in updates.items():
            if field_name not in header_map:
                continue
            converted_updates[header_map[field_name]] = value

        return self.update_row_by_key(
            sheet_name=sheet_name,
            key_col_idx=header_map[key_name],
            key_value=key_value,
            updates=converted_updates
        )

    def replace_full_row_by_key(
        self,
        sheet_name: str,
        key_col_idx: int,
        key_value: Any,
        new_row: Sequence[Any]
    ) -> bool:
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)

        headers = self.get_sheet_headers(sheet_name)
        row_values = list(new_row)[:len(headers)]
        row_values += [""] * (len(headers) - len(row_values))

        key_value = norm_text(key_value)
        for excel_row in range(2, ws.max_row + 1):
            if norm_text(ws.cell(excel_row, key_col_idx + 1).value) == key_value:
                for i, value in enumerate(row_values, start=1):
                    ws.cell(excel_row, i, value)
                self.mark_dirty()
                self.invalidate_sheet_cache(sheet_name)
                self._save_or_defer()
                return True
        return False

    def upsert_row(
        self,
        sheet_name: str,
        key_col_idx: int,
        key_value: Any,
        full_row: Sequence[Any]
    ) -> str:
        """
        Returns 'updated' or 'inserted'
        """
        if self.exists(sheet_name, key_col_idx, key_value):
            self.replace_full_row_by_key(sheet_name, key_col_idx, key_value, full_row)
            return "updated"

        self.append_row(sheet_name, full_row)
        return "inserted"

    def upsert_dict(
        self,
        sheet_name: str,
        key_name: str,
        row_dict: Dict[str, Any]
    ) -> str:
        headers = self.get_sheet_headers(sheet_name)
        header_map = self.get_header_index_map(sheet_name)
        if key_name not in header_map:
            raise ValueError(f"Unknown key column '{key_name}' for sheet '{sheet_name}'")

        key_value = row_dict.get(key_name, "")
        row = [row_dict.get(h, "") for h in headers]
        return self.upsert_row(sheet_name, header_map[key_name], key_value, row)

    # --------------------------------------------------------
    # Delete operations
    # --------------------------------------------------------
    def delete_row_by_key(self, sheet_name: str, key_col_idx: int, key_value: Any) -> bool:
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)

        key_value = norm_text(key_value)
        for excel_row in range(2, ws.max_row + 1):
            if norm_text(ws.cell(excel_row, key_col_idx + 1).value) == key_value:
                ws.delete_rows(excel_row, 1)
                self.mark_dirty()
                self.invalidate_sheet_cache(sheet_name)
                self._save_or_defer()
                return True
        return False

    def delete_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any) -> bool:
        header_map = self.get_header_index_map(sheet_name)
        if key_name not in header_map:
            raise ValueError(f"Unknown key column '{key_name}' for sheet '{sheet_name}'")
        return self.delete_row_by_key(sheet_name, header_map[key_name], key_value)

    def delete_rows_where(self, sheet_name: str, predicate: Callable[[Dict[str, Any]], bool]) -> int:
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)
        headers = self.get_sheet_headers(sheet_name)

        rows_to_delete = []
        for excel_row in range(2, ws.max_row + 1):
            row_vals = [ws.cell(excel_row, i + 1).value for i in range(len(headers))]
            row_dict = dict(zip(headers, row_vals))
            if predicate(row_dict):
                rows_to_delete.append(excel_row)

        for excel_row in reversed(rows_to_delete):
            ws.delete_rows(excel_row, 1)

        if rows_to_delete:
            self.mark_dirty()
            self.invalidate_sheet_cache(sheet_name)
            self._save_or_defer()

        return len(rows_to_delete)

    # --------------------------------------------------------
    # Bulk / utility operations
    # --------------------------------------------------------
    def clear_sheet_data(self, sheet_name: str) -> None:
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)

        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)
            self.mark_dirty()
            self.invalidate_sheet_cache(sheet_name)
            self._save_or_defer()

    def rewrite_sheet_data(self, sheet_name: str, rows: List[Sequence[Any]]) -> None:
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)
        headers = self.get_sheet_headers(sheet_name)

        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)

        for row in rows:
            row_vals = list(row)[:len(headers)]
            row_vals += [""] * (len(headers) - len(row_vals))
            ws.append(row_vals)

        self.mark_dirty()
        self.invalidate_sheet_cache(sheet_name)
        self._save_or_defer()

    def reorder_rows_by_field(
        self,
        sheet_name: str,
        filter_key: str,
        filter_value: Any,
        order_field: str,
        ordered_key_field: str,
        ordered_key_values: List[Any]
    ) -> None:
        """
        Reorders a filtered group of records by assigning order_field = 1..N
        based on ordered_key_values.
        """
        self.ensure_ready()
        wb = self.load_workbook_safe()
        ws = self.get_sheet(wb, sheet_name)
        header_map = self.get_header_index_map(sheet_name)

        if filter_key not in header_map or order_field not in header_map or ordered_key_field not in header_map:
            raise ValueError("Invalid field name in reorder_rows_by_field")

        filter_idx = header_map[filter_key]
        order_idx = header_map[order_field]
        ordered_key_idx = header_map[ordered_key_field]

        order_lookup = {norm_text(v): i + 1 for i, v in enumerate(ordered_key_values)}

        changed = False
        for excel_row in range(2, ws.max_row + 1):
            if norm_text(ws.cell(excel_row, filter_idx + 1).value) != norm_text(filter_value):
                continue

            key_val = norm_text(ws.cell(excel_row, ordered_key_idx + 1).value)
            if key_val in order_lookup:
                ws.cell(excel_row, order_idx + 1, order_lookup[key_val])
                changed = True

        if changed:
            self.mark_dirty()
            self.invalidate_sheet_cache(sheet_name)
            self._save_or_defer()

    # --------------------------------------------------------
    # Relationship helpers
    # --------------------------------------------------------
    def get_rows_by_owner(self, sheet_name: str, owner_type: str, owner_code: str) -> List[Dict[str, Any]]:
        return self.filter_dicts(
            sheet_name,
            lambda r: norm_text(r.get("OwnerType")) == norm_text(owner_type)
            and norm_text(r.get("OwnerCode")) == norm_text(owner_code)
        )

    def delete_rows_by_owner(self, sheet_name: str, owner_type: str, owner_code: str) -> int:
        return self.delete_rows_where(
            sheet_name,
            lambda r: norm_text(r.get("OwnerType")) == norm_text(owner_type)
            and norm_text(r.get("OwnerCode")) == norm_text(owner_code)
        )

    # --------------------------------------------------------
    # Docs folder helpers
    # --------------------------------------------------------
    def get_module_docs_folder(self, module_code: str) -> Path:
        AppConfig.ensure_directories()
        folder = AppConfig.MODULE_DOCS_DIR / safe_name(module_code)
        folder.mkdir(parents=True, exist_ok=True)
        return folder


class PostgresRepository:
    """
    Repository implementation backed by PostgreSQL.

    It mirrors the ExcelRepository method surface closely so the
    service layer can switch backends without major changes.
    """

    backend_name = "postgres"

    def __init__(self):
        load_project_env()
        self._sheet_cache: Dict[str, List[Dict[str, Any]]] = {}
        self._batch_depth = 0
        self._batch_conn = None

    @property
    def connection_label(self) -> str:
        settings = get_database_settings()
        return f"PostgreSQL ({settings['host']} / {settings['dbname']})"

    def invalidate_sheet_cache(self, sheet_name: Optional[str] = None) -> None:
        if sheet_name is None:
            self._sheet_cache = {}
        else:
            self._sheet_cache.pop(sheet_name, None)

    def reload_cache(self) -> None:
        self.invalidate_sheet_cache()

    def mark_dirty(self) -> None:
        return None

    def has_workbook(self) -> bool:
        return True

    def require_workbook_path(self) -> str:
        raise ValueError("Workbook paths are not used in remote data mode.")

    def ensure_ready(self) -> None:
        load_project_env()
        AppConfig.ensure_directories()

    def get_sheet_headers(self, sheet_name: str) -> List[str]:
        headers = AppConfig.SHEET_HEADERS.get(sheet_name)
        if not headers:
            raise ValueError(f"Unknown sheet name: {sheet_name}")
        return headers

    def get_header_index_map(self, sheet_name: str) -> Dict[str, int]:
        return {header: idx for idx, header in enumerate(self.get_sheet_headers(sheet_name))}

    def get_table_name(self, sheet_name: str) -> str:
        table_name = SHEET_TABLE_NAMES.get(sheet_name)
        if not table_name:
            raise ValueError(f"Unknown sheet name: {sheet_name}")
        return table_name

    def get_column_names(self, sheet_name: str) -> List[str]:
        return [header_to_column_name(header) for header in self.get_sheet_headers(sheet_name)]

    def get_primary_key(self, sheet_name: str) -> tuple[str, str]:
        headers = self.get_sheet_headers(sheet_name)
        columns = self.get_column_names(sheet_name)
        return headers[0], columns[0]

    def _get_connection(self):
        self.ensure_ready()
        if self._batch_conn is not None:
            return self._batch_conn, False
        return connect_postgres(), True

    def _finish_connection(self, conn, owns_connection: bool, success: bool) -> None:
        if not owns_connection:
            return
        try:
            if success:
                conn.commit()
            else:
                conn.rollback()
        finally:
            conn.close()

    @contextmanager
    def batch_update(self):
        self.ensure_ready()
        created_conn = False
        if self._batch_depth == 0:
            self._batch_conn = connect_postgres()
            created_conn = True
        self._batch_depth += 1
        try:
            yield
            if created_conn and self._batch_conn is not None:
                self._batch_conn.commit()
        except Exception:
            if created_conn and self._batch_conn is not None:
                self._batch_conn.rollback()
            raise
        finally:
            self._batch_depth -= 1
            if created_conn and self._batch_conn is not None:
                self._batch_conn.close()
                self._batch_conn = None
                self.invalidate_sheet_cache()

    def _fetch_sheet_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
        sql = get_psycopg_sql()
        table_name = self.get_table_name(sheet_name)
        headers = self.get_sheet_headers(sheet_name)
        columns = self.get_column_names(sheet_name)
        _, primary_key_column = self.get_primary_key(sheet_name)
        query = sql.SQL("SELECT {cols} FROM {table} ORDER BY {pk}").format(
            cols=sql.SQL(", ").join(sql.Identifier(col) for col in columns),
            table=sql.Identifier(table_name),
            pk=sql.Identifier(primary_key_column),
        )

        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(query)
                rows = cur.fetchall()
            success = True
            return [dict(zip(headers, row)) for row in rows]
        finally:
            self._finish_connection(conn, owns_connection, success)

    def list_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
        if sheet_name not in self._sheet_cache:
            self._sheet_cache[sheet_name] = self._fetch_sheet_dicts(sheet_name)
        return [dict(row) for row in self._sheet_cache[sheet_name]]

    def list_rows(self, sheet_name: str) -> List[List[Any]]:
        headers = self.get_sheet_headers(sheet_name)
        return [[row.get(header) for header in headers] for row in self.list_dicts(sheet_name)]

    def filter_rows(self, sheet_name: str, predicate: Callable[[List[Any]], bool]) -> List[List[Any]]:
        return [row for row in self.list_rows(sheet_name) if predicate(row)]

    def filter_dicts(self, sheet_name: str, predicate: Callable[[Dict[str, Any]], bool]) -> List[Dict[str, Any]]:
        return [row for row in self.list_dicts(sheet_name) if predicate(row)]

    def find_row(self, sheet_name: str, key_col_idx: int, key_value: Any) -> Optional[List[Any]]:
        key_value = norm_text(key_value)
        for row in self.list_rows(sheet_name):
            if key_col_idx < len(row) and norm_text(row[key_col_idx]) == key_value:
                return row
        return None

    def read_sheet_as_rows(self, sheet_name: str) -> List[List[Any]]:
        return self.list_rows(sheet_name)

    def read_sheet_as_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
        return self.list_dicts(sheet_name)

    def find_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any) -> Optional[Dict[str, Any]]:
        header_map = self.get_header_index_map(sheet_name)
        row = self.find_row(sheet_name, header_map[key_name], key_value)
        if row is None:
            return None
        headers = self.get_sheet_headers(sheet_name)
        return dict(zip(headers, row))

    def find_row_index(self, sheet_name: str, key_col_idx: int, key_value: Any) -> Optional[int]:
        for idx, row in enumerate(self.list_rows(sheet_name), start=2):
            if key_col_idx < len(row) and norm_text(row[key_col_idx]) == norm_text(key_value):
                return idx
        return None

    def exists(self, sheet_name: str, key_col_idx: int, key_value: Any) -> bool:
        return self.find_row_index(sheet_name, key_col_idx, key_value) is not None

    def append_row(self, sheet_name: str, values: Sequence[Any]) -> None:
        sql = get_psycopg_sql()
        headers = self.get_sheet_headers(sheet_name)
        row_values = list(values)[:len(headers)] + [None] * (len(headers) - len(values))
        table_name = self.get_table_name(sheet_name)
        columns = self.get_column_names(sheet_name)
        query = sql.SQL("INSERT INTO {table} ({cols}) VALUES ({vals})").format(
            table=sql.Identifier(table_name),
            cols=sql.SQL(", ").join(sql.Identifier(col) for col in columns),
            vals=sql.SQL(", ").join(sql.Placeholder() for _ in columns),
        )

        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(query, row_values)
            success = True
            self.invalidate_sheet_cache(sheet_name)
        finally:
            self._finish_connection(conn, owns_connection, success)

    def append_dict(self, sheet_name: str, row_dict: Dict[str, Any]) -> None:
        headers = self.get_sheet_headers(sheet_name)
        self.append_row(sheet_name, [row_dict.get(header) for header in headers])

    def update_row_by_key(self, sheet_name: str, key_col_idx: int, key_value: Any, updates: Dict[int, Any]) -> bool:
        sql = get_psycopg_sql()
        if not updates:
            return False
        headers = self.get_sheet_headers(sheet_name)
        columns = self.get_column_names(sheet_name)
        table_name = self.get_table_name(sheet_name)
        key_column = columns[key_col_idx]

        set_parts = []
        params: list[Any] = []
        for col_idx, value in updates.items():
            set_parts.append(sql.SQL("{} = {}").format(sql.Identifier(columns[col_idx]), sql.Placeholder()))
            params.append(value)
        params.append(key_value)

        query = sql.SQL("UPDATE {table} SET {updates} WHERE {key_col} = {key_val}").format(
            table=sql.Identifier(table_name),
            updates=sql.SQL(", ").join(set_parts),
            key_col=sql.Identifier(key_column),
            key_val=sql.Placeholder(),
        )

        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(query, params)
                updated = cur.rowcount > 0
            success = True
            if updated:
                self.invalidate_sheet_cache(sheet_name)
            return updated
        finally:
            self._finish_connection(conn, owns_connection, success)

    def update_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any, updates: Dict[str, Any]) -> bool:
        header_map = self.get_header_index_map(sheet_name)
        converted = {header_map[field]: value for field, value in updates.items() if field in header_map}
        return self.update_row_by_key(sheet_name, header_map[key_name], key_value, converted)

    def replace_full_row_by_key(self, sheet_name: str, key_col_idx: int, key_value: Any, new_row: Sequence[Any]) -> bool:
        headers = self.get_sheet_headers(sheet_name)
        values = list(new_row)[:len(headers)] + [None] * (len(headers) - len(new_row))
        return self.update_row_by_key(sheet_name, key_col_idx, key_value, {idx: value for idx, value in enumerate(values)})

    def upsert_row(self, sheet_name: str, key_col_idx: int, key_value: Any, full_row: Sequence[Any]) -> str:
        sql = get_psycopg_sql()
        headers = self.get_sheet_headers(sheet_name)
        row_values = list(full_row)[:len(headers)] + [None] * (len(headers) - len(full_row))
        table_name = self.get_table_name(sheet_name)
        columns = self.get_column_names(sheet_name)
        key_column = columns[key_col_idx]
        update_columns = [col for idx, col in enumerate(columns) if idx != key_col_idx]

        query = sql.SQL(
            "INSERT INTO {table} ({cols}) VALUES ({vals}) "
            "ON CONFLICT ({key_col}) DO UPDATE SET {updates}"
        ).format(
            table=sql.Identifier(table_name),
            cols=sql.SQL(", ").join(sql.Identifier(col) for col in columns),
            vals=sql.SQL(", ").join(sql.Placeholder() for _ in columns),
            key_col=sql.Identifier(key_column),
            updates=sql.SQL(", ").join(
                sql.SQL("{col} = EXCLUDED.{col}").format(col=sql.Identifier(col)) for col in update_columns
            ),
        )

        existed = self.exists(sheet_name, key_col_idx, key_value)
        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(query, row_values)
            success = True
            self.invalidate_sheet_cache(sheet_name)
        finally:
            self._finish_connection(conn, owns_connection, success)
        return "updated" if existed else "inserted"

    def upsert_dict(self, sheet_name: str, key_name: str, row_dict: Dict[str, Any]) -> str:
        headers = self.get_sheet_headers(sheet_name)
        header_map = self.get_header_index_map(sheet_name)
        row = [row_dict.get(header) for header in headers]
        return self.upsert_row(sheet_name, header_map[key_name], row_dict.get(key_name), row)

    def delete_row_by_key(self, sheet_name: str, key_col_idx: int, key_value: Any) -> bool:
        sql = get_psycopg_sql()
        table_name = self.get_table_name(sheet_name)
        key_column = self.get_column_names(sheet_name)[key_col_idx]
        query = sql.SQL("DELETE FROM {table} WHERE {key_col} = {key_val}").format(
            table=sql.Identifier(table_name),
            key_col=sql.Identifier(key_column),
            key_val=sql.Placeholder(),
        )

        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(query, [key_value])
                deleted = cur.rowcount > 0
            success = True
            if deleted:
                self.invalidate_sheet_cache(sheet_name)
            return deleted
        finally:
            self._finish_connection(conn, owns_connection, success)

    def delete_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any) -> bool:
        header_map = self.get_header_index_map(sheet_name)
        return self.delete_row_by_key(sheet_name, header_map[key_name], key_value)

    def delete_rows_where(self, sheet_name: str, predicate: Callable[[Dict[str, Any]], bool]) -> int:
        rows = self.filter_dicts(sheet_name, predicate)
        if not rows:
            return 0
        key_header, _ = self.get_primary_key(sheet_name)
        deleted = 0
        for row in rows:
            if self.delete_row_by_key_name(sheet_name, key_header, row.get(key_header)):
                deleted += 1
        return deleted

    def clear_sheet_data(self, sheet_name: str) -> None:
        sql = get_psycopg_sql()
        table_name = self.get_table_name(sheet_name)
        query = sql.SQL("TRUNCATE TABLE {table}").format(table=sql.Identifier(table_name))
        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(query)
            success = True
            self.invalidate_sheet_cache(sheet_name)
        finally:
            self._finish_connection(conn, owns_connection, success)

    def rewrite_sheet_data(self, sheet_name: str, rows: List[Sequence[Any]]) -> None:
        with self.batch_update():
            self.clear_sheet_data(sheet_name)
            for row in rows:
                self.append_row(sheet_name, row)

    def reorder_rows_by_field(
        self,
        sheet_name: str,
        filter_key: str,
        filter_value: Any,
        order_field: str,
        ordered_key_field: str,
        ordered_key_values: List[Any],
    ) -> None:
        with self.batch_update():
            for order, ordered_key in enumerate(ordered_key_values, start=1):
                rows = self.filter_dicts(
                    sheet_name,
                    lambda r: norm_text(r.get(filter_key)) == norm_text(filter_value)
                    and norm_text(r.get(ordered_key_field)) == norm_text(ordered_key),
                )
                for row in rows:
                    key_header, _ = self.get_primary_key(sheet_name)
                    self.update_row_by_key_name(sheet_name, key_header, row.get(key_header), {order_field: order})

    def get_rows_by_owner(self, sheet_name: str, owner_type: str, owner_code: str) -> List[Dict[str, Any]]:
        return self.filter_dicts(
            sheet_name,
            lambda r: norm_text(r.get("OwnerType")) == norm_text(owner_type)
            and norm_text(r.get("OwnerCode")) == norm_text(owner_code),
        )

    def delete_rows_by_owner(self, sheet_name: str, owner_type: str, owner_code: str) -> int:
        return self.delete_rows_where(
            sheet_name,
            lambda r: norm_text(r.get("OwnerType")) == norm_text(owner_type)
            and norm_text(r.get("OwnerCode")) == norm_text(owner_code),
        )

    def get_module_docs_folder(self, module_code: str) -> Path:
        AppConfig.ensure_directories()
        folder = AppConfig.MODULE_DOCS_DIR / safe_name(module_code)
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def get_product_docs_folder(self, product_code: str) -> Path:
        AppConfig.ensure_directories()
        folder = AppConfig.PRODUCT_DOCS_DIR / safe_name(product_code)
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def get_project_docs_folder(self, project_code: str) -> Path:
        AppConfig.ensure_directories()
        folder = AppConfig.PROJECT_DOCS_DIR / safe_name(project_code)
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def save_document_blob(
        self,
        sheet_name: str,
        record_id: str,
        file_name: str,
        source_file_path: str,
        created_on: Optional[str] = None,
        updated_on: Optional[str] = None,
    ) -> str:
        src = Path(source_file_path)
        data = src.read_bytes()
        content_type = mimetypes.guess_type(src.name)[0] or "application/octet-stream"
        timestamp = updated_on or created_on or now_str()
        query = """
            INSERT INTO document_blobs
                (sheet_name, record_id, file_name, original_path, content_type, file_data, file_size, created_on, updated_on)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (sheet_name, record_id)
            DO UPDATE SET
                file_name = EXCLUDED.file_name,
                original_path = EXCLUDED.original_path,
                content_type = EXCLUDED.content_type,
                file_data = EXCLUDED.file_data,
                file_size = EXCLUDED.file_size,
                updated_on = EXCLUDED.updated_on
        """

        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(
                    query,
                    [
                        sheet_name,
                        record_id,
                        file_name,
                        str(src),
                        content_type,
                        data,
                        len(data),
                        created_on or timestamp,
                        timestamp,
                    ],
                )
            success = True
        finally:
            self._finish_connection(conn, owns_connection, success)

        return f"db://{sheet_name}/{record_id}/{safe_name(file_name)}"

    def get_document_blob(self, sheet_name: str, record_id: str) -> Optional[Dict[str, Any]]:
        query = """
            SELECT file_name, original_path, content_type, file_data, file_size, created_on, updated_on
            FROM document_blobs
            WHERE sheet_name = %s AND record_id = %s
        """
        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(query, [sheet_name, record_id])
                row = cur.fetchone()
            success = True
            if row is None:
                return None
            keys = ["file_name", "original_path", "content_type", "file_data", "file_size", "created_on", "updated_on"]
            return dict(zip(keys, row))
        finally:
            self._finish_connection(conn, owns_connection, success)

    def materialize_document_blob(self, sheet_name: str, record_id: str, fallback_path: str = "") -> Path:
        blob = self.get_document_blob(sheet_name, record_id)
        if blob is None:
            fallback = Path(fallback_path) if fallback_path else None
            if fallback and fallback.exists():
                return fallback
            raise FileNotFoundError(f"No attachment stored for {sheet_name}:{record_id}")

        AppConfig.ensure_directories()
        folder = AppConfig.TEMP_DIR / "db_attachments" / safe_name(sheet_name)
        folder.mkdir(parents=True, exist_ok=True)

        file_name = blob.get("file_name") or Path(fallback_path).name or record_id
        target = folder / safe_name(file_name)
        data = blob.get("file_data") or b""
        if not target.exists() or target.stat().st_size != len(data):
            target.write_bytes(data)
        return target

    def delete_document_blob(self, sheet_name: str, record_id: str) -> None:
        query = "DELETE FROM document_blobs WHERE sheet_name = %s AND record_id = %s"
        conn, owns_connection = self._get_connection()
        success = False
        try:
            with conn.cursor() as cur:
                cur.execute(query, [sheet_name, record_id])
            success = True
        finally:
            self._finish_connection(conn, owns_connection, success)


class ApiRepository:
    """
    Repository implementation backed directly by PostgREST.

    It preserves the repository surface expected by the desktop app
    while translating operations into PostgREST table requests.
    """

    backend_name = "api"

    def __init__(self):
        load_project_env()
        self.base_url = os.getenv("ERP_API_BASE_URL", "http://127.0.0.1:3000").rstrip("/")
        self.api_prefix = os.getenv("ERP_API_PREFIX", "").strip()
        self._sheet_cache: Dict[str, List[Dict[str, Any]]] = {}

    @property
    def connection_label(self) -> str:
        return f"PostgREST ({self.base_url}{self.api_prefix})"

    def get_table_name(self, sheet_name: str) -> str:
        table_name = SHEET_TABLE_NAMES.get(sheet_name)
        if not table_name:
            raise ValueError(f"Unknown sheet name: {sheet_name}")
        return table_name

    def get_column_names(self, sheet_name: str) -> List[str]:
        return [header_to_column_name(header) for header in self.get_sheet_headers(sheet_name)]

    def _api_url(self, path: str, query: Optional[Dict[str, Any]] = None) -> str:
        base = f"{self.base_url}{self.api_prefix}{path}"
        if not query:
            return base
        filtered = {k: v for k, v in query.items() if v is not None}
        if not filtered:
            return base
        return f"{base}?{urlencode(filtered, doseq=True)}"

    def _request(
        self,
        method: str,
        path: str,
        payload: Any = None,
        query: Optional[Dict[str, Any]] = None,
        headers: Optional[Dict[str, str]] = None,
        timeout: int = 30,
    ) -> tuple[Any, Any]:
        data = None
        request_headers = {"Accept": "application/json"}
        if headers:
            request_headers.update(headers)
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")
            request_headers["Content-Type"] = "application/json"
        request = Request(self._api_url(path, query), data=data, headers=request_headers, method=method)
        try:
            with urlopen(request, timeout=timeout) as response:
                raw = response.read()
                response_headers = response.headers
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace").strip()
            if detail:
                raise RuntimeError(f"{method} {self._api_url(path, query)} failed: HTTP {exc.code} {detail}") from exc
            raise RuntimeError(f"{method} {self._api_url(path, query)} failed: HTTP {exc.code}") from exc
        if not raw:
            return None, response_headers
        return json.loads(raw.decode("utf-8")), response_headers

    def _request_json(
        self,
        method: str,
        path: str,
        payload: Any = None,
        query: Optional[Dict[str, Any]] = None,
        headers: Optional[Dict[str, str]] = None,
        timeout: int = 30,
    ) -> Any:
        payload_data, _ = self._request(method, path, payload=payload, query=query, headers=headers, timeout=timeout)
        return payload_data

    def _request_single_json(
        self,
        method: str,
        path: str,
        payload: Any = None,
        query: Optional[Dict[str, Any]] = None,
        headers: Optional[Dict[str, str]] = None,
    ) -> Optional[Dict[str, Any]]:
        request_headers = {
            "Accept": "application/vnd.pgrst.object+json",
        }
        if headers:
            request_headers.update(headers)
        try:
            return self._request_json(method, path, payload=payload, query=query, headers=request_headers)
        except RuntimeError as exc:
            if "HTTP 406" in str(exc) or "HTTP 404" in str(exc):
                return None
            raise

    def _build_select(self, sheet_name: str) -> str:
        headers = self.get_sheet_headers(sheet_name)
        columns = self.get_column_names(sheet_name)
        return ",".join(f"{header}:{column}" for header, column in zip(headers, columns))

    def _blob_select(self) -> str:
        return ",".join(
            [
                "sheet_name",
                "record_id",
                "file_name",
                "original_path",
                "content_type",
                "file_data",
                "file_size",
                "created_on",
                "updated_on",
            ]
        )

    def _encode_filter_value(self, value: Any) -> str:
        if value is None or value == "":
            return "is.null"
        return f"eq.{value}"

    def _map_row_to_payload(self, sheet_name: str, row_dict: Dict[str, Any]) -> Dict[str, Any]:
        headers = self.get_sheet_headers(sheet_name)
        columns = self.get_column_names(sheet_name)
        payload: Dict[str, Any] = {}
        for header, column in zip(headers, columns):
            if header in row_dict:
                payload[column] = row_dict.get(header)
        return payload

    def _decode_bytea(self, value: Any) -> bytes:
        if value in (None, ""):
            return b""
        if isinstance(value, bytes):
            return value
        text = str(value)
        if text.startswith("\\x"):
            return bytes.fromhex(text[2:])
        return text.encode("utf-8")

    def _fetch_table_rows(self, sheet_name: str, filters: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        table_name = self.get_table_name(sheet_name)
        _, primary_key_column = self.get_primary_key(sheet_name)
        query = {
            "select": self._build_select(sheet_name),
            "order": primary_key_column,
        }
        if filters:
            for header, value in filters.items():
                query[header_to_column_name(header)] = self._encode_filter_value(value)
        rows = self._request_json("GET", f"/{table_name}", query=query)
        return rows or []

    def invalidate_sheet_cache(self, sheet_name: Optional[str] = None) -> None:
        if sheet_name is None:
            self._sheet_cache = {}
        else:
            self._sheet_cache.pop(sheet_name, None)

    def reload_cache(self) -> None:
        self.invalidate_sheet_cache()

    def mark_dirty(self) -> None:
        return None

    def has_workbook(self) -> bool:
        return True

    def require_workbook_path(self) -> str:
        raise ValueError("Workbook paths are not used in online mode.")

    def ensure_ready(self) -> None:
        load_project_env()
        AppConfig.ensure_directories()

    @contextmanager
    def batch_update(self):
        self.ensure_ready()
        try:
            yield
        finally:
            self.invalidate_sheet_cache()

    def get_sheet_headers(self, sheet_name: str) -> List[str]:
        headers = AppConfig.SHEET_HEADERS.get(sheet_name)
        if not headers:
            raise ValueError(f"Unknown sheet name: {sheet_name}")
        return headers

    def get_header_index_map(self, sheet_name: str) -> Dict[str, int]:
        return {header: idx for idx, header in enumerate(self.get_sheet_headers(sheet_name))}

    def get_primary_key(self, sheet_name: str) -> tuple[str, str]:
        headers = self.get_sheet_headers(sheet_name)
        return headers[0], header_to_column_name(headers[0])

    def list_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
        if sheet_name not in self._sheet_cache:
            rows = self._fetch_table_rows(sheet_name)
            self._sheet_cache[sheet_name] = rows or []
        return [dict(row) for row in self._sheet_cache[sheet_name]]

    def list_rows(self, sheet_name: str) -> List[List[Any]]:
        headers = self.get_sheet_headers(sheet_name)
        return [[row.get(header) for header in headers] for row in self.list_dicts(sheet_name)]

    def filter_rows(self, sheet_name: str, predicate: Callable[[List[Any]], bool]) -> List[List[Any]]:
        return [row for row in self.list_rows(sheet_name) if predicate(row)]

    def filter_dicts(self, sheet_name: str, predicate: Callable[[Dict[str, Any]], bool]) -> List[Dict[str, Any]]:
        return [row for row in self.list_dicts(sheet_name) if predicate(row)]

    def find_row(self, sheet_name: str, key_col_idx: int, key_value: Any) -> Optional[List[Any]]:
        key_header = self.get_sheet_headers(sheet_name)[key_col_idx]
        record = self.find_row_by_key_name(sheet_name, key_header, key_value)
        if record is None:
            return None
        headers = self.get_sheet_headers(sheet_name)
        return [record.get(header) for header in headers]

    def read_sheet_as_rows(self, sheet_name: str) -> List[List[Any]]:
        return self.list_rows(sheet_name)

    def read_sheet_as_dicts(self, sheet_name: str) -> List[Dict[str, Any]]:
        return self.list_dicts(sheet_name)

    def find_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any) -> Optional[Dict[str, Any]]:
        table_name = self.get_table_name(sheet_name)
        row = self._request_single_json(
            "GET",
            f"/{table_name}",
            query={
                "select": self._build_select(sheet_name),
                header_to_column_name(key_name): self._encode_filter_value(key_value),
            },
        )
        return row

    def find_row_index(self, sheet_name: str, key_col_idx: int, key_value: Any) -> Optional[int]:
        for idx, row in enumerate(self.list_rows(sheet_name), start=2):
            if key_col_idx < len(row) and norm_text(row[key_col_idx]) == norm_text(key_value):
                return idx
        return None

    def exists(self, sheet_name: str, key_col_idx: int, key_value: Any) -> bool:
        return self.find_row_index(sheet_name, key_col_idx, key_value) is not None

    def append_row(self, sheet_name: str, values: Sequence[Any]) -> None:
        headers = self.get_sheet_headers(sheet_name)
        row = {header: value for header, value in zip(headers, list(values)[:len(headers)])}
        self.append_dict(sheet_name, row)

    def append_dict(self, sheet_name: str, row_dict: Dict[str, Any]) -> None:
        table_name = self.get_table_name(sheet_name)
        payload = self._map_row_to_payload(sheet_name, row_dict)
        self._request_json("POST", f"/{table_name}", payload=payload, headers={"Prefer": "return=minimal"})
        self.invalidate_sheet_cache(sheet_name)

    def update_row_by_key(self, sheet_name: str, key_col_idx: int, key_value: Any, updates: Dict[int, Any]) -> bool:
        headers = self.get_sheet_headers(sheet_name)
        payload = {headers[idx]: value for idx, value in updates.items() if idx < len(headers)}
        key_name = headers[key_col_idx]
        result = self._request_json(
            "PATCH",
            f"/{self.get_table_name(sheet_name)}",
            payload=self._map_row_to_payload(sheet_name, payload),
            query={header_to_column_name(key_name): self._encode_filter_value(key_value)},
            headers={"Prefer": "return=representation"},
        )
        self.invalidate_sheet_cache(sheet_name)
        return bool(result)

    def update_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any, updates: Dict[str, Any]) -> bool:
        result = self._request_json(
            "PATCH",
            f"/{self.get_table_name(sheet_name)}",
            payload=self._map_row_to_payload(sheet_name, updates),
            query={header_to_column_name(key_name): self._encode_filter_value(key_value)},
            headers={"Prefer": "return=representation"},
        )
        self.invalidate_sheet_cache(sheet_name)
        return bool(result)

    def replace_full_row_by_key(self, sheet_name: str, key_col_idx: int, key_value: Any, new_row: Sequence[Any]) -> bool:
        headers = self.get_sheet_headers(sheet_name)
        row = {header: value for header, value in zip(headers, list(new_row)[:len(headers)])}
        key_name = headers[key_col_idx]
        row[key_name] = key_value
        status = self.upsert_dict(sheet_name, key_name, row)
        return status in {"updated", "inserted"}

    def upsert_row(self, sheet_name: str, key_col_idx: int, key_value: Any, full_row: Sequence[Any]) -> str:
        headers = self.get_sheet_headers(sheet_name)
        key_name = headers[key_col_idx]
        row = {header: value for header, value in zip(headers, list(full_row)[:len(headers)])}
        row[key_name] = key_value
        return self.upsert_dict(sheet_name, key_name, row)

    def upsert_dict(self, sheet_name: str, key_name: str, row_dict: Dict[str, Any]) -> str:
        key_value = row_dict.get(key_name)
        existed = self.find_row_by_key_name(sheet_name, key_name, key_value) is not None
        self._request_json(
            "POST",
            f"/{self.get_table_name(sheet_name)}",
            payload=self._map_row_to_payload(sheet_name, row_dict),
            query={"on_conflict": header_to_column_name(key_name)},
            headers={
                "Prefer": "resolution=merge-duplicates,return=representation",
            },
        )
        self.invalidate_sheet_cache(sheet_name)
        return "updated" if existed else "inserted"

    def delete_row_by_key(self, sheet_name: str, key_col_idx: int, key_value: Any) -> bool:
        key_name = self.get_sheet_headers(sheet_name)[key_col_idx]
        return self.delete_row_by_key_name(sheet_name, key_name, key_value)

    def delete_row_by_key_name(self, sheet_name: str, key_name: str, key_value: Any) -> bool:
        result = self._request_json(
            "DELETE",
            f"/{self.get_table_name(sheet_name)}",
            query={header_to_column_name(key_name): self._encode_filter_value(key_value)},
            headers={"Prefer": "return=representation"},
        )
        self.invalidate_sheet_cache(sheet_name)
        return bool(result)

    def delete_rows_where(self, sheet_name: str, predicate: Callable[[Dict[str, Any]], bool]) -> int:
        rows = self.filter_dicts(sheet_name, predicate)
        if not rows:
            return 0
        key_header, _ = self.get_primary_key(sheet_name)
        deleted = 0
        for row in rows:
            if self.delete_row_by_key_name(sheet_name, key_header, row.get(key_header)):
                deleted += 1
        return deleted

    def clear_sheet_data(self, sheet_name: str) -> None:
        self._request_json(
            "DELETE",
            f"/{self.get_table_name(sheet_name)}",
            query={self.get_primary_key(sheet_name)[1]: "not.is.null"},
            headers={"Prefer": "return=minimal"},
        )
        self.invalidate_sheet_cache(sheet_name)

    def rewrite_sheet_data(self, sheet_name: str, rows: List[Sequence[Any]]) -> None:
        headers = self.get_sheet_headers(sheet_name)
        payload = [{header: value for header, value in zip(headers, list(row)[:len(headers)])} for row in rows]
        self.clear_sheet_data(sheet_name)
        for row in payload:
            self.append_dict(sheet_name, row)

    def reorder_rows_by_field(
        self,
        sheet_name: str,
        filter_key: str,
        filter_value: Any,
        order_field: str,
        ordered_key_field: str,
        ordered_key_values: List[Any],
    ) -> None:
        with self.batch_update():
            for order, ordered_key in enumerate(ordered_key_values, start=1):
                rows = self.filter_dicts(
                    sheet_name,
                    lambda r: norm_text(r.get(filter_key)) == norm_text(filter_value)
                    and norm_text(r.get(ordered_key_field)) == norm_text(ordered_key),
                )
                key_header, _ = self.get_primary_key(sheet_name)
                for row in rows:
                    self.update_row_by_key_name(sheet_name, key_header, row.get(key_header), {order_field: order})

    def get_rows_by_owner(self, sheet_name: str, owner_type: str, owner_code: str) -> List[Dict[str, Any]]:
        return self.filter_dicts(
            sheet_name,
            lambda r: norm_text(r.get("OwnerType")) == norm_text(owner_type)
            and norm_text(r.get("OwnerCode")) == norm_text(owner_code),
        )

    def delete_rows_by_owner(self, sheet_name: str, owner_type: str, owner_code: str) -> int:
        return self.delete_rows_where(
            sheet_name,
            lambda r: norm_text(r.get("OwnerType")) == norm_text(owner_type)
            and norm_text(r.get("OwnerCode")) == norm_text(owner_code),
        )

    def get_module_docs_folder(self, module_code: str) -> Path:
        folder = AppConfig.TEMP_DIR / "api_uploads" / safe_name(module_code)
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def get_product_docs_folder(self, product_code: str) -> Path:
        folder = AppConfig.TEMP_DIR / "api_uploads" / safe_name(product_code)
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def get_project_docs_folder(self, project_code: str) -> Path:
        folder = AppConfig.TEMP_DIR / "api_uploads" / safe_name(project_code)
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    def save_document_blob(
        self,
        sheet_name: str,
        record_id: str,
        file_name: str,
        source_file_path: str,
        created_on: Optional[str] = None,
        updated_on: Optional[str] = None,
    ) -> str:
        src = Path(source_file_path)
        content_type = mimetypes.guess_type(src.name)[0] or "application/octet-stream"
        data = src.read_bytes()
        payload = {
            "sheet_name": sheet_name,
            "record_id": record_id,
            "file_name": file_name,
            "original_path": str(src),
            "content_type": content_type,
            "file_data": "\\x" + data.hex(),
            "file_size": len(data),
            "created_on": created_on or updated_on or now_str(),
            "updated_on": updated_on or created_on or now_str(),
        }
        self._request_json(
            "POST",
            "/document_blobs",
            payload=payload,
            query={"on_conflict": "sheet_name,record_id"},
            headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
        )
        return f"db://{sheet_name}/{record_id}/{safe_name(file_name)}"

    def get_document_blob(self, sheet_name: str, record_id: str) -> Optional[Dict[str, Any]]:
        row = self._request_single_json(
            "GET",
            "/document_blobs",
            query={
                "select": self._blob_select(),
                "sheet_name": self._encode_filter_value(sheet_name),
                "record_id": self._encode_filter_value(record_id),
            },
        )
        if row is None:
            return None
        return {
            "file_name": row.get("file_name") or f"{record_id}.bin",
            "original_path": row.get("original_path") or "",
            "content_type": row.get("content_type") or "application/octet-stream",
            "file_data": self._decode_bytea(row.get("file_data")),
            "file_size": row.get("file_size") or 0,
            "created_on": row.get("created_on") or "",
            "updated_on": row.get("updated_on") or "",
        }

    def materialize_document_blob(self, sheet_name: str, record_id: str, fallback_path: str = "") -> Path:
        blob = self.get_document_blob(sheet_name, record_id)
        if blob is None:
            fallback = Path(fallback_path) if fallback_path else None
            if fallback and fallback.exists():
                return fallback
            raise FileNotFoundError(f"No attachment stored for {sheet_name}:{record_id}")
        AppConfig.ensure_directories()
        folder = AppConfig.TEMP_DIR / "api_attachments" / safe_name(sheet_name)
        folder.mkdir(parents=True, exist_ok=True)
        target = folder / safe_name(blob["file_name"])
        target.write_bytes(blob["file_data"])
        return target

    def delete_document_blob(self, sheet_name: str, record_id: str) -> None:
        self._request_json(
            "DELETE",
            "/document_blobs",
            query={
                "sheet_name": self._encode_filter_value(sheet_name),
                "record_id": self._encode_filter_value(record_id),
            },
            headers={"Prefer": "return=minimal"},
        )


# ============================================================
# Convenience facade
# ============================================================

class WorkbookManager:
    """
    Small helper wrapper around ExcelRepository for UI code.

    Lets pages do simple checks like:
    - has workbook?
    - create workbook
    - open workbook
    """

    def __init__(self, backend: Optional[str] = None):
        load_project_env()
        selected_backend = norm_text(backend or os.getenv("ERP_DATA_BACKEND") or "local").lower()
        if selected_backend == "online":
            self.repo = ApiRepository()
            self.backend_name = "online"
        elif selected_backend == "local":
            self.repo = ExcelRepository()
            self.backend_name = "local"
        else:
            raise ValueError("ERP_DATA_BACKEND must be either 'online' or 'local'.")

    @property
    def workbook_path(self) -> Optional[str]:
        if not self.uses_local():
            return self.repo.connection_label
        return self.repo.workbook_path

    def uses_online(self) -> bool:
        return self.backend_name == "online"

    def uses_local(self) -> bool:
        return self.backend_name == "local"

    def set_workbook_path(self, path: str) -> None:
        if not self.uses_local():
            raise ValueError("Workbook paths are only used in local mode.")
        self.repo.set_workbook_path(path)

    def has_workbook(self) -> bool:
        return self.repo.has_workbook()

    def create_workbook(self, path: str) -> None:
        if not self.uses_local():
            raise ValueError("Workbook creation is only available in local mode.")
        AppConfig.ensure_directories()
        WorkbookSchema.create_new_workbook(path)
        self.repo.set_workbook_path(path)

    def open_workbook(self, path: str) -> None:
        if not self.uses_local():
            raise ValueError("Workbook selection is only available in local mode.")
        AppConfig.ensure_directories()
        WorkbookSchema.ensure_workbook_structure(path)
        self.repo.set_workbook_path(path)
